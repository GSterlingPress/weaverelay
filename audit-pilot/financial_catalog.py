from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import csv, re, subprocess
from typing import Any

from source_registry import SourceRegistry

FINANCIAL_WORDS = {
    "A": ("rate", "fee", "price", "cost", "hour", "quantity", "markup", "multiplier", "credit", "invoice", "payroll", "retainage"),
    "B": ("shall", "must", "limited", "maximum", "adjusted", "factor", "percent", "amendment", "supersede", "schedule", "exhibit", "rate"),
    "C": ("invoice", "amount", "hours", "qty", "quantity", "rate", "subtotal", "total", "credit", "receipt", "timesheet", "payroll"),
}

REFERENCE_PATTERNS = (
    re.compile(r"(?i)\b(?:exhibit|schedule|appendix|attachment)\s+[A-Z0-9.-]+"),
    re.compile(r"(?i)\b(?:amendment|change order|modification)\s+(?:no\.?\s*)?[A-Z0-9.-]+"),
)
NUMBER_RE = re.compile(r"(?<!\w)(?:\$\s*)?-?\d[\d,]*(?:\.\d+)?\s*%?")


def _pdf_text(path: Path) -> list[str]:
    try:
        p = subprocess.run(["pdftotext", "-layout", str(path), "-"], capture_output=True, text=True, timeout=120)
        text = p.stdout if p.returncode == 0 else ""
    except Exception:
        text = ""
    if not text:
        return []
    # pdftotext uses form-feed between pages.
    return text.split("\f")


def _xlsx_text(path: Path) -> list[tuple[str, int, str]]:
    out: list[tuple[str, int, str]] = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for rno, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    out.append((ws.title, rno, " | ".join(vals)))
        wb.close()
    except Exception:
        pass
    return out


def _csv_text(path: Path) -> list[tuple[int, str]]:
    out=[]
    try:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
            for rno, row in enumerate(csv.reader(fh), 1):
                out.append((rno, " | ".join(row)))
    except Exception:
        pass
    return out


def _interesting(text: str, profile: str) -> bool:
    low=text.lower()
    return any(word in low for word in FINANCIAL_WORDS[profile]) or bool(NUMBER_RE.search(text))


def build_financial_catalog(registry: SourceRegistry, profile: str) -> dict[str, Any]:
    """Build one analyzer's own catalog from raw registered sources.

    Profiles deliberately use different relevance lenses. They share only source
    IDs; no analyzer receives another analyzer's catalog or conclusions.
    """
    if profile not in {"A", "B", "C"}:
        raise ValueError(profile)
    facts=[]; references=[]; unreadable=[]
    for doc in registry.manifest()["documents"]:
        path=Path(doc["path"]); doc_id=doc["doc_id"]
        ext=path.suffix.lower(); found_any=False
        records=[]
        if ext==".pdf":
            records=[("page", i, text) for i,text in enumerate(_pdf_text(path),1)]
        elif ext in {".xlsx", ".xlsm"}:
            records=[("sheetrow", (sheet,row), text) for sheet,row,text in _xlsx_text(path)]
        elif ext==".csv":
            records=[("row", row, text) for row,text in _csv_text(path)]
        elif ext in {".txt", ".md", ".eml"}:
            try:
                records=[("line", i, line) for i,line in enumerate(path.read_text(encoding="utf-8",errors="replace").splitlines(),1)]
            except Exception:
                records=[]
        elif ext in {".jpg",".jpeg",".png",".webp",".tif",".tiff",".bmp"}:
            # OCR is analyzer-specific; use tesseract without modifying original.
            try:
                p=subprocess.run(["tesseract",str(path),"stdout"],capture_output=True,text=True,timeout=120)
                records=[("region","ocr",p.stdout)] if p.returncode==0 else []
            except Exception:
                records=[]

        for kind, loc, text in records:
            if not text or not _interesting(text,profile):
                continue
            found_any=True
            if kind=="page": locator=registry.locator(doc_id,page=int(loc))
            elif kind=="sheetrow": locator=registry.locator(doc_id,sheet=str(loc[0]),row=loc[1])
            elif kind=="row": locator=registry.locator(doc_id,row=loc)
            else: locator=registry.locator(doc_id,region=str(loc))
            snippet=" ".join(text.split())[:1200]
            numbers=NUMBER_RE.findall(snippet)
            facts.append({"locator":locator,"document":doc_id,"role":doc["role"],"snippet":snippet,"numbers":numbers})
            for pat in REFERENCE_PATTERNS:
                for match in pat.findall(snippet):
                    references.append({"locator":locator,"reference":match})
        if not found_any and ext not in {".zip"}:
            unreadable.append(doc_id)

    # Completeness is intentionally conservative: referenced items are surfaced,
    # not auto-declared missing without a deterministic name match.
    names=" ".join(d["original_name"].lower() for d in registry.manifest()["documents"])
    reference_checks=[]
    for ref in references:
        token=ref["reference"].lower()
        significant=[p for p in re.split(r"\s+",token) if len(p)>1 and p not in {"exhibit","schedule","appendix","attachment","amendment","change","order","modification","no."}]
        present=bool(significant) and all(p.strip(".,") in names for p in significant)
        reference_checks.append({**ref,"appears_present_by_filename":present})

    return {
        "analyzer":profile,
        "catalog_version":"zero-dave-catalog-v1",
        "facts":facts,
        "references":reference_checks,
        "unreadable_documents":unreadable,
        "fact_count":len(facts),
    }
