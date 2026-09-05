from __future__ import annotations

from pathlib import Path
import csv, re, subprocess
from typing import Any

LABOR_PATTERNS=(
    re.compile(r"(?i)overall\s+(?:multiplier|factor)\s+(?:of\s+)?([0-9]+(?:\.[0-9]+)?)"),
    re.compile(r"(?i)(?:multiplier|factor)\s+(?:is\s+|of\s+)?([0-9]+(?:\.[0-9]+)?).{0,140}?(?:actual|raw).{0,80}?(?:hourly|payroll|labor|salary)"),
)
FEE_PATTERNS=(
    re.compile(r"(?i)subcontract(?:or)?\s+(?:fee|markup|mark-up)[^.]{0,120}?(?:shall\s+not\s+exceed|limited\s+to|maximum|cap(?:ped)?\s+at|should\s+only\s+be)\s*([0-9]+(?:\.[0-9]+)?)\s*(?:%|percent)"),
    re.compile(r"(?i)(?:shall\s+not\s+exceed|limited\s+to|maximum|cap(?:ped)?\s+at)\s*([0-9]+(?:\.[0-9]+)?)\s*(?:%|percent)[^.]{0,100}?subcontract"),
)


def _text(path: str) -> str:
    p=Path(path)
    if p.suffix.lower()=='.pdf':
        try:
            r=subprocess.run(['pdftotext','-layout',str(p),'-'],capture_output=True,text=True,timeout=120)
            return r.stdout if r.returncode==0 else ''
        except Exception:return ''
    try:return p.read_text(encoding='utf-8',errors='replace')
    except Exception:return ''


def _rows(path: str) -> list[dict[str,Any]]:
    p=Path(path); ext=p.suffix.lower()
    if ext=='.csv':
        with p.open('r',encoding='utf-8-sig',errors='replace',newline='') as f:
            return [dict(r) for r in csv.DictReader(f)]
    if ext in {'.xlsx','.xlsm'}:
        try:
            from openpyxl import load_workbook
            wb=load_workbook(p,read_only=True,data_only=True); out=[]
            for ws in wb.worksheets:
                it=ws.iter_rows(values_only=True)
                headers=[str(x or '').strip() for x in next(it,())]
                for row in it: out.append({headers[i]:row[i] for i in range(min(len(headers),len(row))) if headers[i]})
            wb.close(); return out
        except Exception:return []
    return []


def _f(v:Any)->float|None:
    try:
        if v is None or str(v).strip()=='':return None
        return float(str(v).replace('$','').replace(',','').replace('%','').strip())
    except Exception:return None


def _first(row:dict,*keys):
    low={str(k).lower():v for k,v in row.items()}
    for key in keys:
        if key.lower() in low and low[key.lower()] not in (None,''):return low[key.lower()]
    return None


def _extract_rules(contract_text:str)->dict:
    mult=None; mtxt=None
    for pat in LABOR_PATTERNS:
        m=pat.search(contract_text)
        if m: mult=float(m.group(1)); mtxt=m.group(0); break
    cap=None; ctxt=None
    for pat in FEE_PATTERNS:
        m=pat.search(contract_text)
        if m: cap=float(m.group(1)); ctxt=m.group(0); break
    return {'labor_multiplier':mult,'labor_text':mtxt,'subcontract_cap_pct':cap,'subcontract_text':ctxt}


def run_analyzer_b(*,contract:str,invoice:str,field:str|None,evidence:list[str],registry)->dict:
    """Independent contract-first financial engine.

    B never calls audit_engine. It extracts governing numeric rules directly from
    contract text, then applies those rules to normalized invoice/evidence rows.
    """
    rules=_extract_rules(_text(contract)); inv_rows=_rows(invoice)
    ev_rows=[]
    for p in ([field] if field else [])+list(evidence):
        if p: ev_rows.extend(_rows(p))
    contract_id=registry.doc_id_for_path(contract); invoice_id_doc=registry.doc_id_for_path(invoice)
    ev_doc_ids=[registry.doc_id_for_path(p) for p in evidence if registry.doc_id_for_path(p)]
    findings=[]

    # Evidence index is deliberately keyed contract-first by stable invoice identity
    # and then classification/description, never by amount alone.
    for i,row in enumerate(inv_rows,2):
        iid=str(_first(row,'invoice_id','invoice','pay_app') or '').strip()
        desc=str(_first(row,'description','rate_key','classification','type') or '').strip()
        hours=_f(_first(row,'hours','qty','quantity'))
        billed_rate=_f(_first(row,'rate','billed_rate','unit_rate'))
        if rules['labor_multiplier'] is not None and hours is not None and billed_rate is not None:
            matches=[]
            for er in ev_rows:
                ei=str(_first(er,'invoice_id','invoice','pay_app') or '').strip()
                ed=str(_first(er,'classification','rate_key','description','type') or '').strip()
                if iid and ei==iid and (not desc or not ed or desc.lower() in ed.lower() or ed.lower() in desc.lower()):
                    pr=_f(_first(er,'rate','payroll_rate','hourly_rate','actual_rate'))
                    if pr is not None: matches.append((pr,er))
            if len(matches)==1:
                payroll=matches[0][0]; allowed=payroll*rules['labor_multiplier']; amount=max(0.0,(billed_rate-allowed)*hours)
                if amount>0.004:
                    findings.append({
                        'invoice_id':iid,'description':desc,'code':'RATE_MISMATCH','status':'OVERBILLED','confidence':'HIGH','amount':round(amount,2),
                        'rule':rules['labor_multiplier'],'rule_source':registry.locator(contract_id,region='LABOR-RULE'),
                        'invoice_source':registry.locator(invoice_id_doc,row=i),'evidence_source':ev_doc_ids,
                        'formula':'(billed_rate - (payroll_rate * multiplier)) * hours',
                        'inputs':{'billed_rate':round(billed_rate,6),'payroll_rate':round(payroll,6),'multiplier':rules['labor_multiplier'],'hours':round(hours,6)},
                        '_zero_dave_trace':{'rule_kind':'labor_multiplier','rule_value':rules['labor_multiplier'],'invoice_id':iid,'description':desc,'formula_id':'LABOR_RATE_DELTA','inputs':{'billed_rate':round(billed_rate,6),'payroll_rate':round(payroll,6),'multiplier':rules['labor_multiplier'],'hours':round(hours,6)}}
                    })

        base=_f(_first(row,'subcontractor_base_cost','subcontractor_cost','base_cost'))
        billed_pct=_f(_first(row,'markup_pct','fee_pct','fee_percent','markup_percent'))
        if rules['subcontract_cap_pct'] is not None and base is not None and billed_pct is not None and billed_pct>rules['subcontract_cap_pct']:
            amount=base*((billed_pct-rules['subcontract_cap_pct'])/100.0)
            findings.append({
                'invoice_id':iid,'description':desc,'code':'SUBCONTRACTOR_MARKUP_EXCEEDED','status':'OVERBILLED','confidence':'HIGH','amount':round(amount,2),
                'rule':rules['subcontract_cap_pct'],'rule_source':registry.locator(contract_id,region='SUBCONTRACT-FEE-RULE'),
                'invoice_source':registry.locator(invoice_id_doc,row=i),'evidence_source':[],
                'formula':'base_cost * ((billed_pct - cap_pct) / 100)',
                'inputs':{'base_cost':round(base,6),'billed_pct':round(billed_pct,6),'cap_pct':rules['subcontract_cap_pct']},
                '_zero_dave_trace':{'rule_kind':'subcontract_cap_pct','rule_value':rules['subcontract_cap_pct'],'invoice_id':iid,'description':desc,'formula_id':'SUBCONTRACT_FEE_DELTA','inputs':{'base_cost':round(base,6),'billed_pct':round(billed_pct,6),'cap_pct':rules['subcontract_cap_pct']}}
            })
    total=round(sum(float(f['amount']) for f in findings),2)
    return {'analyzer':'B','engine':'contract-first-v1','rules':rules,'findings':findings,'totals':{'overbilled':total,'unsupported':0.0},'unknown':[]}
