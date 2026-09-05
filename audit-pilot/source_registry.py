from __future__ import annotations

from dataclasses import dataclass, asdict
from hashlib import sha256
from pathlib import Path
from typing import Iterable, Optional
import mimetypes


@dataclass(frozen=True)
class SourceDocument:
    doc_id: str
    role: str
    original_name: str
    path: str
    extension: str
    media_type: str
    size_bytes: int
    sha256: str
    page_count: Optional[int] = None
    sheet_names: tuple[str, ...] = ()

    def to_dict(self) -> dict:
        data = asdict(self)
        data["sheet_names"] = list(self.sheet_names)
        return data


class SourceRegistry:
    """Neutral immutable numbering layer shared by all analyzers.

    The registry records where evidence is. It intentionally does not decide what
    the evidence means. Analyzer A/B/C each build their own financial catalog on
    top of these identifiers.
    """

    def __init__(self, audit_id: str):
        self.audit_id = audit_id
        self._documents: list[SourceDocument] = []
        self._path_to_id: dict[str, str] = {}

    @staticmethod
    def _fingerprint(path: Path) -> str:
        h = sha256()
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()

    @staticmethod
    def _pdf_page_count(path: Path) -> Optional[int]:
        if path.suffix.lower() != ".pdf":
            return None
        try:
            from pypdf import PdfReader
            return len(PdfReader(str(path)).pages)
        except Exception:
            try:
                import fitz  # PyMuPDF
                with fitz.open(str(path)) as doc:
                    return len(doc)
            except Exception:
                return None

    @staticmethod
    def _sheet_names(path: Path) -> tuple[str, ...]:
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return ()
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=False)
            names = tuple(wb.sheetnames)
            wb.close()
            return names
        except Exception:
            return ()

    def add(self, path: str | Path, role: str, original_name: Optional[str] = None) -> SourceDocument:
        p = Path(path).resolve()
        key = str(p)
        if key in self._path_to_id:
            return self.by_id(self._path_to_id[key])
        doc_id = f"DOC-{len(self._documents) + 1:04d}"
        media_type = mimetypes.guess_type(original_name or p.name)[0] or "application/octet-stream"
        item = SourceDocument(
            doc_id=doc_id,
            role=role,
            original_name=original_name or p.name,
            path=key,
            extension=p.suffix.lower(),
            media_type=media_type,
            size_bytes=p.stat().st_size,
            sha256=self._fingerprint(p),
            page_count=self._pdf_page_count(p),
            sheet_names=self._sheet_names(p),
        )
        self._documents.append(item)
        self._path_to_id[key] = doc_id
        return item

    def add_many(self, items: Iterable[tuple[str | Path, str, Optional[str]]]) -> list[SourceDocument]:
        return [self.add(path, role, name) for path, role, name in items]

    def by_id(self, doc_id: str) -> SourceDocument:
        for item in self._documents:
            if item.doc_id == doc_id:
                return item
        raise KeyError(doc_id)

    def doc_id_for_path(self, path: str | Path) -> Optional[str]:
        return self._path_to_id.get(str(Path(path).resolve()))

    @staticmethod
    def locator(
        doc_id: str,
        *,
        page: Optional[int] = None,
        region: Optional[str] = None,
        table: Optional[str] = None,
        row: Optional[str | int] = None,
        cell: Optional[str] = None,
        sheet: Optional[str] = None,
    ) -> str:
        parts = [doc_id]
        if page is not None:
            parts.append(f"P-{int(page):04d}")
        if sheet:
            parts.append(f"S-{sheet}")
        if region:
            parts.append(f"R-{region}")
        if table:
            parts.append(f"T-{table}")
        if row is not None:
            parts.append(f"ROW-{row}")
        if cell:
            parts.append(f"CELL-{cell}")
        return "/".join(parts)

    def manifest(self) -> dict:
        return {
            "audit_id": self.audit_id,
            "numbering_policy": "immutable-append-only",
            "documents": [item.to_dict() for item in self._documents],
        }
