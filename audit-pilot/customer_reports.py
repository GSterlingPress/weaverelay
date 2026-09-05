from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile, ZIP_DEFLATED
import csv, html, subprocess

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether

PALETTE = [colors.HexColor(x) for x in ("#2F6BFF","#0F9D78","#9C5CFF","#D97706","#C2415D","#1976A3","#7A5C00","#4B5563")]


def _styles():
    s=getSampleStyleSheet()
    s.add(ParagraphStyle(name="CenterTitle",parent=s["Title"],alignment=TA_CENTER,spaceAfter=18))
    s.add(ParagraphStyle(name="SmallGray",parent=s["BodyText"],fontSize=8,textColor=colors.HexColor("#555555"),leading=10))
    s.add(ParagraphStyle(name="Finding",parent=s["Heading2"],spaceBefore=14,spaceAfter=8))
    return s


def _money(v):
    try:return f"${float(v or 0):,.2f}"
    except Exception:return "$0.00"


def build_findings_pdf(result: dict, output: str | Path) -> Path:
    out=Path(output); styles=_styles()
    story=[Paragraph("Audit Findings Report",styles["CenterTitle"]),Paragraph(f"Audit ID: <b>{html.escape(result['audit_id'])}</b>",styles["BodyText"]),Spacer(1,10)]
    c=result.get("consensus",{})
    summary=Table([
        ["Verified findings",_money(c.get("verified_total"))],
        ["Unresolved potential (excluded)",_money(c.get("unresolved_potential_total"))],
        ["Documents cataloged",str(len(result.get("source_registry",{}).get("documents",[])))],
    ],colWidths=[3.2*inch,2.2*inch])
    summary.setStyle(TableStyle([("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D0D5DD")),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F4F6F8")),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("PADDING",(0,0),(-1,-1),8)]))
    story += [summary,Spacer(1,14),Paragraph("SHOW ME THIS DOLLAR",styles["Heading2"]),Paragraph("Every released dollar is paired with its analyzer traces and canonical source IDs. Unresolved disagreements are disclosed and excluded from the verified total.",styles["BodyText"])]

    for idx,f in enumerate(c.get("findings",[])):
        color=PALETTE[idx%len(PALETTE)]
        fid=html.escape(f.get("finding_id","")); state=html.escape(f.get("state",""))
        story.append(Paragraph(f"<font color='{color.hexval()}'>{fid}</font> — {state} — {_money(f.get('verified_amount') or f.get('potential_amount'))}",styles["Finding"]))
        story.append(Paragraph(f"Materiality: <b>{html.escape(str(f.get('materiality_band','')))}</b> · Reconciliation budget: {f.get('reconciliation_budget',0)} · First divergence: {html.escape(str(f.get('divergence_stage') or 'none'))}",styles["BodyText"]))
        rows=[["Analyzer","Status","Amount","Contract source","Invoice source","Evidence source"]]
        for t in f.get("traces",[]):
            def short(v):
                x=str(v or "")
                return html.escape(x[:85]+("…" if len(x)>85 else ""))
            rows.append([t.get("analyzer",""),t.get("status",""),_money(t.get("amount")),Paragraph(short(t.get("rule_source")),styles["SmallGray"]),Paragraph(short(t.get("invoice_source")),styles["SmallGray"]),Paragraph(short(t.get("evidence_source")),styles["SmallGray"])])
        table=Table(rows,repeatRows=1,colWidths=[.55*inch,.8*inch,.8*inch,1.35*inch,1.35*inch,1.35*inch])
        table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),color),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#D0D5DD")),("VALIGN",(0,0),(-1,-1),"TOP"),("FONTSIZE",(0,0),(-1,-1),7),("PADDING",(0,0),(-1,-1),4)]))
        story.append(table)
        story.append(Paragraph(html.escape(f.get("note","")),styles["SmallGray"]))

    story += [PageBreak(),Paragraph("Important use notice",styles["Heading2"]),Paragraph("This report is automated contract/invoice analysis and decision support. It may contain OCR, extraction, interpretation, matching, or calculation errors. Unresolved findings are not included in the verified total. Customers should verify material findings before relying on them or pursuing recovery. Final customer terms and limitation-of-liability language require attorney review before commercial release.",styles["BodyText"])]
    SimpleDocTemplate(str(out),pagesize=letter,rightMargin=36,leftMargin=36,topMargin=40,bottomMargin=40,title="Audit Findings Report").build(story)
    return out


def _pdf_images(path: Path, temp: Path):
    prefix=temp/(path.stem+"-page")
    p=subprocess.run(["pdftoppm","-png","-r","110",str(path),str(prefix)],capture_output=True,text=True,timeout=600)
    if p.returncode!=0:return []
    return sorted(temp.glob(path.stem+"-page-*.png"))


def _text_preview_story(path: Path, styles):
    ext=path.suffix.lower(); story=[]
    if ext==".csv":
        with path.open("r",encoding="utf-8-sig",errors="replace",newline="") as fh:
            for i,row in enumerate(csv.reader(fh),1):
                story.append(Paragraph(f"<b>Row {i}:</b> {html.escape(' | '.join(row))}",styles["SmallGray"]))
    elif ext in {".txt",".md",".eml"}:
        for i,line in enumerate(path.read_text(encoding="utf-8",errors="replace").splitlines(),1):
            story.append(Paragraph(f"<b>Line {i}:</b> {html.escape(line)}",styles["SmallGray"]))
    elif ext in {".xlsx",".xlsm"}:
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True)
        for ws in wb.worksheets:
            story += [Paragraph(f"Sheet: {html.escape(ws.title)}",styles["Heading3"])]
            for rno,row in enumerate(ws.iter_rows(values_only=True),1):
                vals=[str(v) for v in row if v is not None]
                if vals:story.append(Paragraph(f"<b>Row {rno}:</b> {html.escape(' | '.join(vals))}",styles["SmallGray"]))
        wb.close()
    elif ext==".msg":
        try:
            import extract_msg
            m=extract_msg.Message(str(path)); text=f"From: {m.sender}\nTo: {m.to}\nSubject: {m.subject}\n\n{m.body or ''}"
            for i,line in enumerate(text.splitlines(),1):story.append(Paragraph(f"<b>Line {i}:</b> {html.escape(line)}",styles["SmallGray"]))
            m.close()
        except Exception as e:story.append(Paragraph(f"Message rendering unavailable: {html.escape(str(e))}",styles["SmallGray"]))
    return story


def build_evidence_catalog_pdf(result: dict, output: str | Path) -> Path:
    out=Path(output); styles=_styles(); story=[Paragraph("Audit Evidence & Analytical Catalog",styles["CenterTitle"]),Paragraph(f"Audit ID: <b>{html.escape(result['audit_id'])}</b>",styles["BodyText"]),Paragraph("SHOW ME THE SOURCE — every submitted item is accounted for below. The original evidence is preserved separately; these are audit renderings and do not alter the originals.",styles["BodyText"]),PageBreak()]
    with TemporaryDirectory() as td:
        temp=Path(td)
        for doc in result.get("source_registry",{}).get("documents",[]):
            p=Path(doc["path"]); did=doc["doc_id"]
            story += [Paragraph(f"{did} — {html.escape(doc['original_name'])}",styles["Heading2"]),Paragraph(f"Role: {doc['role']} · Type: {doc['media_type']} · Bytes: {doc['size_bytes']:,}<br/>SHA-256: {doc['sha256']}",styles["SmallGray"])]
            ext=p.suffix.lower()
            rendered=False
            if ext==".pdf":
                imgs=_pdf_images(p,temp)
                for n,img in enumerate(imgs,1):
                    story.append(Paragraph(f"{did} / P-{n:04d}",styles["SmallGray"]))
                    story.append(Image(str(img),width=6.7*inch,height=8.7*inch,kind="proportional"))
                    rendered=True
                    if n<len(imgs):story.append(PageBreak())
            elif ext in {".jpg",".jpeg",".png",".webp",".tif",".tiff",".bmp",".heic"}:
                source=p
                if ext==".heic":
                    try:
                        from PIL import Image as PILImage
                        import pillow_heif
                        pillow_heif.register_heif_opener(); im=PILImage.open(p); source=temp/(p.stem+".png"); im.save(source)
                    except Exception: source=p
                try:
                    story.append(Image(str(source),width=6.7*inch,height=8.5*inch,kind="proportional"));rendered=True
                except Exception: pass
            else:
                blocks=_text_preview_story(p,styles)
                if blocks:story.extend(blocks);rendered=True
            if not rendered:
                story.append(Paragraph("Audit could not create a faithful visual rendering for this format. The original file remains preserved in the Original Evidence Archive and is explicitly disclosed as unrendered rather than silently omitted.",styles["BodyText"]))
            story.append(PageBreak())
    SimpleDocTemplate(str(out),pagesize=letter,rightMargin=28,leftMargin=28,topMargin=30,bottomMargin=30,title="Audit Evidence & Analytical Catalog").build(story)
    return out


def build_original_evidence_archive(result: dict, output: str | Path) -> Path:
    out=Path(output)
    with ZipFile(out,"w",ZIP_DEFLATED) as z:
        for doc in result.get("source_registry",{}).get("documents",[]):
            p=Path(doc["path"])
            safe=Path(doc["original_name"]).name
            z.write(p,arcname=f"{doc['doc_id']}__{safe}")
    return out


def build_customer_package(result: dict, work: str | Path) -> dict:
    work=Path(work)
    findings=build_findings_pdf(result,work/"Audit-Findings-Report.pdf")
    evidence=build_evidence_catalog_pdf(result,work/"Audit-Evidence-Analytical-Catalog.pdf")
    originals=build_original_evidence_archive(result,work/"Audit-Original-Evidence.zip")
    return {"findings_pdf":findings.name,"evidence_catalog_pdf":evidence.name,"original_evidence_archive":originals.name}
