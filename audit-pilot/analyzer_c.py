from __future__ import annotations

from pathlib import Path
import csv, re, subprocess
from typing import Any

SENTENCE_SPLIT=re.compile(r"(?<=[.!?])\s+|\n+")
PCT_RE=re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*(?:%|percent)",re.I)
NUM_RE=re.compile(r"([0-9]+(?:\.[0-9]+)?)")


def _contract_sentences(path:str)->list[str]:
    p=Path(path)
    if p.suffix.lower()=='.pdf':
        try:
            r=subprocess.run(['pdftotext','-raw',str(p),'-'],capture_output=True,text=True,timeout=120)
            text=r.stdout if r.returncode==0 else ''
        except Exception:text=''
    else:
        try:text=p.read_text(encoding='utf-8',errors='replace')
        except Exception:text=''
    return [" ".join(s.split()) for s in SENTENCE_SPLIT.split(text) if s.strip()]


def _tabular(path:str)->list[dict[str,Any]]:
    p=Path(path); ext=p.suffix.lower()
    if ext=='.csv':
        with p.open('r',encoding='utf-8-sig',errors='replace',newline='') as f:return [dict(x) for x in csv.DictReader(f)]
    if ext in {'.xlsx','.xlsm'}:
        try:
            from openpyxl import load_workbook
            wb=load_workbook(p,read_only=True,data_only=True); out=[]
            for ws in wb.worksheets:
                rows=ws.iter_rows(values_only=True); hdr=[str(v or '').strip() for v in next(rows,())]
                for vals in rows:out.append({hdr[i]:vals[i] for i in range(min(len(hdr),len(vals))) if hdr[i]})
            wb.close();return out
        except Exception:return []
    return []


def _num(v:Any)->float|None:
    try:
        if v is None or str(v).strip()=='':return None
        return float(str(v).replace('$','').replace(',','').replace('%','').strip())
    except Exception:return None


def _pick(row:dict,*names):
    low={str(k).strip().lower():v for k,v in row.items()}
    for n in names:
        if n.lower() in low and low[n.lower()] not in (None,''):return low[n.lower()]
    return None


def _score_labor(s:str)->int:
    low=s.lower(); score=0
    if 'multiplier' in low or 'overall factor' in low:score+=4
    if 'actual' in low or 'raw' in low:score+=2
    if any(w in low for w in ('hourly','payroll','labor','salary')):score+=2
    if any(w in low for w in ('shall','applied','adjusted','billing')):score+=1
    if 'draft' in low and 'not executed' in low:score-=10
    return score


def _score_fee(s:str)->int:
    low=s.lower(); score=0
    if 'subcontract' in low:score+=4
    if any(w in low for w in ('fee','markup','mark-up')):score+=2
    if any(w in low for w in ('shall not exceed','limited to','maximum','cap','only')):score+=3
    if 'draft' in low and 'not executed' in low:score-=10
    return score


def _infer_rules(sentences:list[str])->dict:
    labor=[]; fee=[]
    for idx,s in enumerate(sentences,1):
        ls=_score_labor(s)
        if ls>=6:
            nums=[float(x) for x in NUM_RE.findall(s)]
            vals=[x for x in nums if 1.0 < x < 10.0]
            if vals:labor.append((ls,idx,vals[-1],s))
        fs=_score_fee(s)
        if fs>=7:
            vals=[float(x) for x in PCT_RE.findall(s)]
            if vals:fee.append((fs,idx,vals[-1],s))
    labor.sort(key=lambda x:(x[0],x[1]),reverse=True); fee.sort(key=lambda x:(x[0],x[1]),reverse=True)
    return {
        'labor_multiplier':labor[0][2] if labor else None,'labor_sentence':labor[0][3] if labor else None,'labor_sentence_no':labor[0][1] if labor else None,
        'subcontract_cap_pct':fee[0][2] if fee else None,'subcontract_sentence':fee[0][3] if fee else None,'subcontract_sentence_no':fee[0][1] if fee else None,
    }


def run_analyzer_c(*,contract:str,invoice:str,field:str|None,evidence:list[str],registry)->dict:
    """Independent ledger-first financial engine.

    C starts from invoice events, independently reconstructs supporting ledger facts,
    then scores contract sentences to determine the applicable numeric constraint.
    It never calls Analyzer A/B or audit_engine.
    """
    invoice_rows=_tabular(invoice)
    supporting=[]
    for p in ([field] if field else [])+list(evidence):
        if p:supporting.extend(_tabular(p))
    rules=_infer_rules(_contract_sentences(contract))
    cdoc=registry.doc_id_for_path(contract); idoc=registry.doc_id_for_path(invoice)
    edocs=[registry.doc_id_for_path(p) for p in evidence if registry.doc_id_for_path(p)]
    findings=[]

    # Ledger-first: create events, then resolve evidence by invoice/date/classification.
    for rowno,row in enumerate(invoice_rows,2):
        iid=str(_pick(row,'invoice_id','invoice','pay_app') or '').strip()
        desc=str(_pick(row,'description','rate_key','classification','type') or '').strip()
        svc=str(_pick(row,'service_date','date') or '').strip()
        hours=_num(_pick(row,'hours','quantity','qty')); billed=_num(_pick(row,'rate','billed_rate','unit_rate'))
        if rules['labor_multiplier'] is not None and hours is not None and billed is not None:
            candidates=[]
            for er in supporting:
                ei=str(_pick(er,'invoice_id','invoice','pay_app') or '').strip(); es=str(_pick(er,'service_date','date') or '').strip()
                ed=str(_pick(er,'classification','rate_key','description','type') or '').strip()
                rate=_num(_pick(er,'rate','payroll_rate','hourly_rate','actual_rate'))
                if rate is None:continue
                score=0
                if iid and ei==iid:score+=5
                if svc and es==svc:score+=2
                if desc and ed and (desc.lower() in ed.lower() or ed.lower() in desc.lower()):score+=3
                if score>=5:candidates.append((score,rate,er))
            candidates.sort(key=lambda x:x[0],reverse=True)
            if candidates and (len(candidates)==1 or candidates[0][0]>candidates[1][0]):
                payroll=candidates[0][1]; allowed=payroll*rules['labor_multiplier']; amt=max(0.0,(billed-allowed)*hours)
                if amt>0.004:
                    findings.append({
                        'invoice_id':iid,'description':desc,'code':'RATE_MISMATCH','status':'OVERBILLED','confidence':'HIGH','amount':round(amt,2),
                        'rule':rules['labor_multiplier'],'rule_source':registry.locator(cdoc,region=f"SENT-{rules['labor_sentence_no'] or 'NA'}"),
                        'invoice_source':registry.locator(idoc,row=rowno),'evidence_source':edocs,
                        'formula':'(billed_rate - (payroll_rate * multiplier)) * hours',
                        'inputs':{'billed_rate':round(billed,6),'payroll_rate':round(payroll,6),'multiplier':rules['labor_multiplier'],'hours':round(hours,6)},
                        '_zero_dave_trace':{'rule_kind':'labor_multiplier','rule_value':rules['labor_multiplier'],'invoice_id':iid,'description':desc,'formula_id':'LABOR_RATE_DELTA','inputs':{'billed_rate':round(billed,6),'payroll_rate':round(payroll,6),'multiplier':rules['labor_multiplier'],'hours':round(hours,6)}}
                    })

        base=_num(_pick(row,'subcontractor_base_cost','subcontractor_cost','base_cost')); pct=_num(_pick(row,'markup_pct','fee_pct','fee_percent','markup_percent'))
        cap=rules['subcontract_cap_pct']
        if cap is not None and base is not None and pct is not None and pct>cap:
            amt=base*((pct-cap)/100.0)
            findings.append({
                'invoice_id':iid,'description':desc,'code':'SUBCONTRACTOR_MARKUP_EXCEEDED','status':'OVERBILLED','confidence':'HIGH','amount':round(amt,2),
                'rule':cap,'rule_source':registry.locator(cdoc,region=f"SENT-{rules['subcontract_sentence_no'] or 'NA'}"),
                'invoice_source':registry.locator(idoc,row=rowno),'evidence_source':[],
                'formula':'base_cost * ((billed_pct - cap_pct) / 100)',
                'inputs':{'base_cost':round(base,6),'billed_pct':round(pct,6),'cap_pct':cap},
                '_zero_dave_trace':{'rule_kind':'subcontract_cap_pct','rule_value':cap,'invoice_id':iid,'description':desc,'formula_id':'SUBCONTRACT_FEE_DELTA','inputs':{'base_cost':round(base,6),'billed_pct':round(pct,6),'cap_pct':cap}}
            })
    total=round(sum(float(f['amount']) for f in findings),2)
    return {'analyzer':'C','engine':'ledger-first-v1','rules':rules,'findings':findings,'totals':{'overbilled':total,'unsupported':0.0},'unknown':[]}
